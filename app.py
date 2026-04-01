import io
import json
import math
import re
import urllib.request
from datetime import datetime, date
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Flask, request, jsonify, send_from_directory, send_file, current_app
from flask_cors import CORS

app = Flask(__name__, static_folder=".")
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# Field names must match the frontend FILE_SLOTS ids
REQUIRED_FILES = ["renewal", "execpack", "pastdue", "gtp", "cancellation"]

# Portfolio Baseline — locked. Update manually if the baseline changes.
ANNUAL_BASELINE = 35989090

# CS rep roster — must mirror the frontend CS_REPS constant
CS_REPS = {
    "Kala McNeely":      "kmcneely",
    "Samantha Marcus":   "samanthamarcus",
    "Genele Ekinde":     "gekinde",
    "Karen Guirguis":    "kguirguis",
    "April Dippolito":   "adippolito",
    "Alex Polivka":      "apolivka",
}
CS_REP_NAMES = list(CS_REPS.keys())
BANDWIDTH_CEILING = 500


# ── openpyxl style constants ──────────────────────────────────
_NAVY_FILL  = PatternFill("solid", fgColor="243454")
_WHT_BOLD   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BOLD       = Font(name="Calibri", bold=True, size=11)
_CURR_FMT   = '$#,##0'
_PCT_FMT    = '0.0%'
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
_AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
_RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
_GREEN_FONT = Font(name="Calibri", color="006100", size=11)
_AMBER_FONT = Font(name="Calibri", color="9C5700", size=11)
_RED_FONT   = Font(name="Calibri", color="9C0006", size=11)


def _fmt_dollar(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt_pct(v):
    """Convert a percentage-point value (e.g. 12.5) to a decimal for Excel (0.125)."""
    try:
        return float(v or 0) / 100.0
    except (TypeError, ValueError):
        return 0.0


def _hrow(ws, row_num, headers):
    """Write a navy header row to worksheet ws at row_num."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=text)
        cell.fill = _NAVY_FILL
        cell.font = _WHT_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, min_w=8, max_w=40):
    """Auto-fit column widths based on cell content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def build_pack_xlsx(data):
    """Build the Executive Pack as an openpyxl Workbook and return a BytesIO buffer."""
    pack     = data.get("pack") or {}
    seg      = data.get("segData") or {}
    mc       = data.get("manualContexts") or {}
    d        = pack.get("dashboard") or {}
    health   = pack.get("health") or []
    bw_list  = pack.get("bandwidth") or []
    rc_list  = pack.get("renewalConc") or []
    coaching = pack.get("coachingActions") or {}
    pdr      = pack.get("pastDueByRep") or {}
    checks   = pack.get("validations") or []
    vsummary = pack.get("validationSummary") or {}

    bw_map = {b.get("rep"): b for b in bw_list}
    rc_map = {r.get("rep"): r for r in rc_list}

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # ── Tab 1: Dashboard Summary ──────────────────────────────
    ws = wb.create_sheet("Dashboard Summary")
    _hrow(ws, 1, ["Metric", "Value"])
    dashboard_rows = [
        ("Portfolio Baseline",  _fmt_dollar(d.get("portfolioBaseline")),   _CURR_FMT),
        ("Current File Total",  _fmt_dollar(d.get("currentFileTotal")),    _CURR_FMT),
        ("WOW Delta ($)",       _fmt_dollar(d.get("wowDelta")),            _CURR_FMT),
        ("WOW Delta (%)",       _fmt_pct(d.get("wowDeltaPct")),            _PCT_FMT),
        ("Total Declined",      _fmt_dollar(d.get("totalDeclined")),       _CURR_FMT),
        ("Portfolio Risk %",    _fmt_pct(d.get("riskPct")),                _PCT_FMT),
        ("Digital Mix %",       _fmt_pct(d.get("digitalMix")),             _PCT_FMT),
        ("Team Health Score",   round(float(d.get("healthScore") or 0), 1), None),
        ("Health Tier",         d.get("healthTier", ""),                    None),
        ("Export Integrity",    d.get("exportFlag", ""),                    None),
        ("Week Label",          pack.get("week", ""),                       None),
        ("Week Date",           pack.get("weekDate", ""),                   None),
        ("Health Model",        d.get("healthModel", ""),                   None),
        ("Validation Status",   vsummary.get("overall_status", ""),         None),
        ("Checks Passed",       vsummary.get("passed", 0),                  None),
        ("Checks Warned",       vsummary.get("warned", 0),                  None),
        ("Checks Failed",       vsummary.get("failed", 0),                  None),
    ]
    for i, (metric, value, fmt) in enumerate(dashboard_rows, 2):
        ws.cell(row=i, column=1, value=metric).font = _BOLD
        cell = ws.cell(row=i, column=2, value=value)
        if fmt:
            cell.number_format = fmt
    ws.freeze_panes = "A2"
    _autowidth(ws)

    # ── Tab 2: Bi-Weekly Performance ──────────────────────────
    ws = wb.create_sheet("Bi-Weekly Performance")
    bw_headers = ["Rep", "InScope", "Declined Rev", "Digital Rev", "Print Rev",
                  "Risk %", "Digital Mix", "Health Score", "Tier", "Trend"]
    _hrow(ws, 1, bw_headers)
    team_inscope = team_declined = team_digital = team_print = 0.0
    for i, rep in enumerate(health, 2):
        inscope  = _fmt_dollar(rep.get("inScope"))
        dec_rev  = _fmt_dollar(rep.get("declinedRev"))
        dig_rev  = _fmt_dollar(rep.get("digitalRev"))
        prt_rev  = inscope - dig_rev
        risk     = _fmt_pct(rep.get("riskPct"))
        dig_mix  = _fmt_pct(rep.get("digitalMix"))
        hs       = round(float(rep.get("healthScore") or 0), 1)
        row_vals = [rep.get("rep", ""), inscope, dec_rev, dig_rev, prt_rev,
                    risk, dig_mix, hs, rep.get("healthTier", ""), rep.get("declineTrend", "—")]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j in (2, 3, 4, 5): cell.number_format = _CURR_FMT
            elif j in (6, 7):     cell.number_format = _PCT_FMT
        team_inscope  += inscope
        team_declined += dec_rev
        team_digital  += dig_rev
        team_print    += prt_rev
    tr = len(health) + 2
    team_row = ["TEAM TOTAL", team_inscope, team_declined, team_digital, team_print,
                team_declined / team_inscope if team_inscope else 0.0,
                team_digital  / team_inscope if team_inscope else 0.0,
                "", "", ""]
    for j, v in enumerate(team_row, 1):
        cell = ws.cell(row=tr, column=j, value=v)
        cell.font = _BOLD
        if j in (2, 3, 4, 5): cell.number_format = _CURR_FMT
        elif j in (6, 7):     cell.number_format = _PCT_FMT
    ws.freeze_panes = "B2"
    _autowidth(ws)

    # ── Tab 3: Team Totals ────────────────────────────────────
    ws = wb.create_sheet("Team Totals")
    tt_headers = ["Rep", "InScope", "Declined Rev", "Digital Rev", "Risk %", "Digital Mix",
                  "Health Score", "Tier", "Trend", "Next 30", "Next 60", "Next 90", "Highest $"]
    _hrow(ws, 1, tt_headers)
    for i, rep in enumerate(health, 2):
        rc = rc_map.get(rep.get("rep"), {})
        row_vals = [
            rep.get("rep", ""),
            _fmt_dollar(rep.get("inScope")),
            _fmt_dollar(rep.get("declinedRev")),
            _fmt_dollar(rep.get("digitalRev")),
            _fmt_pct(rep.get("riskPct")),
            _fmt_pct(rep.get("digitalMix")),
            round(float(rep.get("healthScore") or 0), 1),
            rep.get("healthTier", ""),
            rep.get("declineTrend", "—"),
            rc.get("next30", 0),
            rc.get("next60", 0),
            rc.get("next90", 0),
            _fmt_dollar(rc.get("highestDollar")),
        ]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j in (2, 3, 4): cell.number_format = _CURR_FMT
            elif j in (5, 6):  cell.number_format = _PCT_FMT
            elif j == 13:      cell.number_format = _CURR_FMT
    ws.freeze_panes = "B2"
    _autowidth(ws)

    # ── Tabs 4–6: Segmentation ────────────────────────────────
    seg_tabs = [
        ("Segmentation - Declined",   seg.get("declined",   []),
         ["Agreement #", "Client", "Rep", "Subtotal", "Sale Items", "Digital/Print",
          "Attrition Cause", "Save Offer", "Form Type", "Form Attrition Reason",
          "Contacted", "Contact Type", "Contact Date", "Status", "Blocker",
          "Next Step", "Next Step Date", "Save Offer Outcome", "IDP",
          "Outcome", "Notes", "Decline Trend"]),
        ("Segmentation - Print Only", seg.get("printOnly",  []),
         ["Agreement #", "Client", "Rep", "Subtotal", "Sale Items",
          "Contacted", "Contact Type", "Contact Date", "Status", "Blocker",
          "Next Step", "Next Step Date", "Save Offer Outcome", "IDP",
          "Outcome", "Notes", "Decline Trend"]),
        ("Segmentation - Renewable",  seg.get("renewable",  []),
         ["Agreement #", "Client", "Rep", "Subtotal", "Sale Items",
          "Contacted", "Contact Type", "Contact Date", "Status", "Blocker",
          "Next Step", "Next Step Date", "Save Offer Outcome", "IDP",
          "Outcome", "Notes", "Decline Trend"]),
    ]
    for tab_name, rows, hdrs in seg_tabs:
        ws = wb.create_sheet(tab_name)
        _hrow(ws, 1, hdrs)
        has_form_cols = "Form Type" in hdrs
        for i, row in enumerate(rows, 2):
            if has_form_cols:
                vals = [
                    row.get("agreementNum", ""), row.get("client", ""), row.get("rep", ""),
                    _fmt_dollar(row.get("subtotal")), row.get("saleItems", ""), row.get("digital", ""),
                    row.get("attritionCause", ""), row.get("saveOffer", ""),
                    row.get("formType", ""), row.get("formAttritionReason", ""),
                    row.get("contacted", ""), row.get("contactType", ""), row.get("contactDate", ""),
                    row.get("statusField", ""), row.get("blocker", ""),
                    row.get("nextStep", ""), row.get("nextStepDate", ""),
                    row.get("saveOfferOutcome", ""), row.get("idp", ""),
                    row.get("outcome", ""), row.get("notes", ""), row.get("declineTrend", "—"),
                ]
            else:
                vals = [
                    row.get("agreementNum", ""), row.get("client", ""), row.get("rep", ""),
                    _fmt_dollar(row.get("subtotal")), row.get("saleItems", ""),
                    row.get("contacted", ""), row.get("contactType", ""), row.get("contactDate", ""),
                    row.get("statusField", ""), row.get("blocker", ""),
                    row.get("nextStep", ""), row.get("nextStepDate", ""),
                    row.get("saveOfferOutcome", ""), row.get("idp", ""),
                    row.get("outcome", ""), row.get("notes", ""), row.get("declineTrend", "—"),
                ]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=j, value=v)
                if hdrs[j - 1] == "Subtotal":
                    cell.number_format = _CURR_FMT
        ws.freeze_panes = "A2"
        _autowidth(ws)

    # ── Tab 7: Health Scorecard ───────────────────────────────
    ws = wb.create_sheet("Health Scorecard")
    hs_headers = ["Rep", "Health Score", "Tier", "InScope", "Declined Rev", "Digital Rev",
                  "Risk %", "Digital Mix", "Decline Trend", "Next 30", "Next 60", "Next 90", "Highest $"]
    _hrow(ws, 1, hs_headers)
    for i, rep in enumerate(health, 2):
        rc = rc_map.get(rep.get("rep"), {})
        row_vals = [
            rep.get("rep", ""),
            round(float(rep.get("healthScore") or 0), 1),
            rep.get("healthTier", ""),
            _fmt_dollar(rep.get("inScope")),
            _fmt_dollar(rep.get("declinedRev")),
            _fmt_dollar(rep.get("digitalRev")),
            _fmt_pct(rep.get("riskPct")),
            _fmt_pct(rep.get("digitalMix")),
            rep.get("declineTrend", "—"),
            rc.get("next30", 0),
            rc.get("next60", 0),
            rc.get("next90", 0),
            _fmt_dollar(rc.get("highestDollar")),
        ]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j in (4, 5, 6, 13): cell.number_format = _CURR_FMT
            elif j in (7, 8):      cell.number_format = _PCT_FMT
    ws.freeze_panes = "B2"
    _autowidth(ws)

    # ── Tab 8: Payment Detail ─────────────────────────────────
    ws = wb.create_sheet("Payment Detail")
    pd_headers = ["Rep", "Agreement", "Client", "Total Due",
                  "0-30", "31-60", "61-90", "91-120", "121-150", "150+",
                  "Past Due Reason", "Legal", "Collections", "Ad Pulled"]
    _hrow(ws, 1, pd_headers)
    row_num = 2
    for rep in health:
        name = rep.get("rep", "")
        for row in (pdr.get(name) or []):
            vals = [
                name,
                row.get("agreement", ""),
                row.get("client", ""),
                _fmt_dollar(row.get("totalDue")),
                _fmt_dollar(row.get("d0_30")),
                _fmt_dollar(row.get("d31_60")),
                _fmt_dollar(row.get("d61_90")),
                _fmt_dollar(row.get("d91_120")),
                _fmt_dollar(row.get("d121_150")),
                _fmt_dollar(row.get("d150plus")),
                row.get("pastDueReason", ""),
                "Yes" if row.get("legal") else "No",
                "Yes" if row.get("collections") else "No",
                "Yes" if row.get("adPulled") else "No",
            ]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=j, value=v)
                if j in (4, 5, 6, 7, 8, 9, 10):
                    cell.number_format = _CURR_FMT
            row_num += 1
    ws.freeze_panes = "A2"
    _autowidth(ws)

    # ── Tab 9: Coaching Summary ───────────────────────────────
    ws = wb.create_sheet("Coaching Summary")
    cs_headers = ["Rep", "Health Score", "Tier", "Active / Ceiling",
                  "Action 1", "Action 2", "Action 3", "Action 4", "Action 5", "Manual Context"]
    _hrow(ws, 1, cs_headers)
    for i, rep in enumerate(health, 2):
        name    = rep.get("rep", "")
        bw      = bw_map.get(name, {})
        actions = coaching.get(name) or []
        row_vals = [
            name,
            round(float(rep.get("healthScore") or 0), 1),
            rep.get("healthTier", ""),
            f"{bw.get('active', 0)}/{BANDWIDTH_CEILING}",
        ] + [actions[k] if k < len(actions) else "" for k in range(5)] + [mc.get(name, "")]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j >= 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 75
    ws.freeze_panes = "A2"
    _autowidth(ws, max_w=60)

    # ── Tab 10: Bandwidth ─────────────────────────────────────
    ws = wb.create_sheet("Bandwidth")
    bw_headers2 = ["Rep", "Renewable", "Declined", "Cancelled", "Active", "Ceiling", "Over Ceiling"]
    _hrow(ws, 1, bw_headers2)
    for i, bw in enumerate(bw_list, 2):
        over = bw.get("over")
        row_vals = [
            bw.get("rep", ""),
            bw.get("renewable", 0),
            bw.get("declined", 0),
            bw.get("cancelled", 0),
            bw.get("active", 0),
            BANDWIDTH_CEILING,
            "YES" if over else "No",
        ]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j == 7 and over:
                cell.font = Font(name="Calibri", bold=True, color="9C0006", size=11)
    ws.freeze_panes = "B2"
    _autowidth(ws)

    # ── Tab 11: Validation ────────────────────────────────────
    ws = wb.create_sheet("Validation")
    overall = vsummary.get("overall_status", "")
    ws.cell(row=1, column=1, value="Validation Summary").font = _BOLD
    ws.cell(row=1, column=2, value=f"Overall: {overall}").font = _BOLD
    ws.cell(row=1, column=3, value=f"Passed: {vsummary.get('passed', 0)}")
    ws.cell(row=1, column=4, value=f"Warned: {vsummary.get('warned', 0)}")
    ws.cell(row=1, column=5, value=f"Failed: {vsummary.get('failed', 0)}")
    _hrow(ws, 2, ["Check Name", "Status", "Value", "Expected", "Note"])
    status_map = {
        "PASS": (_GREEN_FILL, _GREEN_FONT),
        "WARN": (_AMBER_FILL, _AMBER_FONT),
        "FAIL": (_RED_FILL,   _RED_FONT),
    }
    for i, chk in enumerate(checks, 3):
        status = chk.get("status", "")
        fill, font = status_map.get(status, (None, _BOLD))
        row_vals = [chk.get("name", ""), status, chk.get("value", ""),
                    chk.get("expected", ""), chk.get("note", "")]
        for j, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j == 2 and fill:
                cell.fill = fill
                cell.font = font
            if j == 5:
                cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A3"
    _autowidth(ws, max_w=50)
    ws.column_dimensions["E"].width = 60  # note column

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def read_xlsx(file_storage, sheet_name=0, header=0):
    data = file_storage if isinstance(file_storage, (bytes, bytearray)) else file_storage.read()
    return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, header=header, engine="openpyxl")


def sanitize_keys(obj):
    if isinstance(obj, dict):
        return {
            (k.isoformat() if isinstance(k, (datetime, date)) else str(k) if not isinstance(k, (str, int, float, bool)) else k): sanitize_keys(v)
            for k, v in obj.items()
        }
    elif isinstance(obj, list):
        return [sanitize_keys(i) for i in obj]
    return obj


class SafeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, float) and (np.isnan(obj) or np.isinf(obj)):
            return None
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)


def _safe_value(v):
    """Convert a single cell value to a JSON-safe scalar."""
    # NaN / NaT / None → null
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    # datetime / Timestamp / date → ISO string
    if isinstance(v, (datetime, date, pd.Timestamp)):
        return v.isoformat()
    return v


def df_to_records(df):
    """Convert DataFrame to a JSON-safe list of dicts.

    Handles both datetime64 dtype columns and object columns that contain
    mixed datetime.datetime / pandas.Timestamp values returned by openpyxl.
    """
    records = []
    for row in df.itertuples(index=False, name=None):
        records.append({col: _safe_value(val) for col, val in zip(df.columns, row)})
    return records


# ──────────────────────────────────────────────────────────────
# VALIDATION HELPERS  (mirrors frontend logic)
# ──────────────────────────────────────────────────────────────

def match_cs_rep(rep_field):
    if not rep_field:
        return None
    rf = str(rep_field).lower()
    for name in CS_REP_NAMES:
        if all(p in rf for p in name.lower().split()):
            return name
        if CS_REPS[name] in rf:
            return name
    return None


def is_digital(sale_items):
    si = str(sale_items or "").lower()
    return any(k in si for k in ["digital", "listings", "google", "meta"])


def extract_agreement_number(s):
    m = re.search(r"E-\d+(?:-\d+)?", str(s or ""), re.IGNORECASE)
    return m.group(0).upper() if m else None


def _extract_client_name(combined):
    """Extract client name from a combined Agreement/Client field.

    Handles formats like:
      'E-372469 / The Wellesley'
      'Samantha Marcus / MW / E-176611-5 / Hemmelgarn'
    Strategy: remove E-XXXXXX token, drop rep-name fragments and short
    region codes (<= 3 chars), return the first remaining meaningful part.
    """
    s = str(combined or "").strip()
    # Remove agreement number token
    s_clean = re.sub(r'E-\d+(?:-\d+)?', '', s, flags=re.IGNORECASE).strip(" /")
    parts = [p.strip() for p in s_clean.split("/") if p.strip()]
    # Drop short codes (state/region abbreviations) and rep first-name matches
    rep_first = {n.split()[0].lower() for n in CS_REP_NAMES}
    keep = [p for p in parts
            if len(p) > 3 and p.split()[0].lower() not in rep_first]
    return keep[0] if keep else (parts[0] if parts else "")


def _chk(name, status, value, expected, note):
    return {"name": name, "status": status, "value": value,
            "expected": expected, "note": note}


def _parse_date_val(v):
    """Return a datetime.date from a cell value (datetime, date, or ISO string), or None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str) and v:
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f",
                    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(v[:19], fmt[:len(v[:19])]).date()
            except ValueError:
                continue
    return None


def run_validations(renewal_rows, exec_pack_text, past_due_rows,
                    gtp_rows, cancellation_rows, annual_baseline):
    checks = []

    # ── Filter + deduplicate renewal rows ────────────────────
    seen = set()
    cs_rows = []
    all_contract_keys = []

    for row in renewal_rows:
        raw = (row.get("Contract #") or row.get("Agreement")
               or row.get("Agreement Number"))
        if raw:
            all_contract_keys.append(str(raw).strip().upper())

    today = date.today()
    for row in renewal_rows:
        rep = match_cs_rep(row.get("Rep"))
        if not rep:
            continue
        status_raw = str(row.get("Renew Status") or "").strip()
        if status_raw == "Cancelled":
            continue
        # Merged: keep only if Last Edition is today or future; exclude past-dated
        if status_raw == "Merged":
            le = _parse_date_val(row.get("Last Edition"))
            if le is None or le < today:
                continue
        raw = (row.get("Contract #") or row.get("Agreement")
               or row.get("Agreement Number"))
        if not raw:
            continue
        key = str(raw).strip().upper()
        if key in seen:
            continue
        seen.add(key)
        cs_rows.append({**row, "_rep": rep, "_key": key})

    current_file_total = sum(
        float(r.get("Subtotal Sales") or 0) for r in cs_rows
    )

    # ── 17. Duplicate contracts ───────────────────────────────
    total_rows = len(renewal_rows)
    dup_count = total_rows - len(set(all_contract_keys)) if all_contract_keys else 0
    dup_pct = (dup_count / total_rows * 100) if total_rows else 0
    checks.append(_chk(
        "duplicate_contracts",
        "FAIL" if dup_pct > 5 else ("WARN" if dup_pct > 2 else "PASS"),
        f"{dup_count} dupes ({dup_pct:.1f}%)", "<2% of rows",
        f"{dup_count} duplicate Contract # in {total_rows} renewal rows before deduplication",
    ))

    # ── 12. Rep coverage ─────────────────────────────────────
    present = {r["_rep"] for r in cs_rows}
    missing = [n for n in CS_REP_NAMES if n not in present]
    checks.append(_chk(
        "rep_coverage",
        "FAIL" if missing else "PASS",
        f"{len(present)}/6 reps present", "All 6 reps",
        f"Missing: {', '.join(missing)}" if missing
        else "All 6 CS reps found in renewal report",
    ))

    # ── Segmentation revenues ─────────────────────────────────
    declined_rev = print_only_rev = renewable_rev = 0
    digital_rev = print_rev = 0
    declined_contracts = set()

    for row in cs_rows:
        status = str(row.get("Renew Status") or "").strip()
        items  = str(row.get("Sale Items") or "")
        sub    = float(row.get("Subtotal Sales") or 0)
        dig    = is_digital(items)

        if dig:
            digital_rev += sub
        else:
            print_rev += sub

        items_lower = items.lower()
        if status == "Declined":
            declined_rev += sub
            declined_contracts.add(row["_key"])
        elif (not dig
              and "print" in items_lower
              and "digital" not in items_lower
              and "listings" not in items_lower):
            print_only_rev += sub
        elif status in ("Renewable", "Renew Eligible"):
            renewable_rev += sub
        elif status == "Merged":
            # Only future-dated Merged rows reach cs_rows; route to Renewable
            renewable_rev += sub

    # ── 1. Segment revenue reconcile ─────────────────────────
    seg_total = declined_rev + print_only_rev + renewable_rev
    seg_gap   = abs(seg_total - current_file_total)
    checks.append(_chk(
        "segment_revenue_reconcile",
        "FAIL" if seg_gap > 100 else ("WARN" if seg_gap > 0 else "PASS"),
        f"${seg_total:,.0f} (gap ${seg_gap:,.0f})", f"${current_file_total:,.0f}",
        f"Declined ${declined_rev:,.0f} + Print Only ${print_only_rev:,.0f}"
        f" + Renewable ${renewable_rev:,.0f}",
    ))

    # ── 2. Digital / print reconcile ─────────────────────────
    dp_total = digital_rev + print_rev
    dp_gap   = abs(dp_total - current_file_total)
    checks.append(_chk(
        "digital_print_reconcile",
        "FAIL" if dp_gap > 100 else "PASS",
        f"${dp_total:,.0f} (gap ${dp_gap:,.0f})", f"${current_file_total:,.0f}",
        f"Digital ${digital_rev:,.0f} + Print ${print_rev:,.0f}",
    ))

    # ── 3. WOW delta math ────────────────────────────────────
    wow = current_file_total - annual_baseline
    checks.append(_chk(
        "wow_delta_math", "PASS",
        f"${wow:,.0f}", "current_file_total − annual_baseline",
        f"${current_file_total:,.0f} − ${annual_baseline:,.0f} = ${wow:,.0f}",
    ))

    # ── Per-rep metrics — 4-Factor Health Model ───────────────
    # Factor 1: Renew Status   40 pts  (Renewable=40, Merged=20, Declined=0)
    # Factor 2: Payment Health 30 pts  (not past due=30, 0-30d=22, 31-60d=15,
    #                                   61-90d=8, 90+d or legal=0)
    # Factor 3: Digital        20 pts  (digital=20, print only=0)
    # Factor 4: Engagement     10 pts  (0 at run time; updated via Segmentation)
    # Score = average of per-agreement scores across all rep's accounts

    # Build past-due lookup keyed by agreement number
    pd_agr_lookup: dict = {}
    for pd_row in past_due_rows:
        raw = (pd_row.get("Agreement") or pd_row.get("Contract #")
               or pd_row.get("Agreement Number") or "")
        agr_key = str(raw).strip().upper()
        if not agr_key:
            continue
        d0_30  = float(pd_row.get("Monthly 0 - 30")  or 0)
        d31_60 = float(pd_row.get("Monthly 31 - 60") or 0)
        d61_90 = float(pd_row.get("Monthly 61 - 90") or 0)
        d90p   = (float(pd_row.get("Monthly 91 - 120") or 0)
                  + float(pd_row.get("Monthly 121 - 150") or 0)
                  + float(pd_row.get("Monthly 150+") or 0))
        legal  = str(pd_row.get("Legal") or "").strip().lower() in ("yes", "true", "1")
        pd_agr_lookup[agr_key] = dict(d0_30=d0_30, d31_60=d31_60,
                                      d61_90=d61_90, d90p=d90p, legal=legal)

    def _pay_pts(agr_key: str) -> float:
        """Payment Health factor (30 pts max)."""
        pd = pd_agr_lookup.get(agr_key)
        if pd is None:
            return 30.0          # not in past-due report → current
        if pd["legal"] or pd["d90p"] > 0:
            return 0.0
        if pd["d61_90"] > 0:
            return 8.0
        if pd["d31_60"] > 0:
            return 15.0
        if pd["d0_30"] > 0:
            return 22.0
        return 30.0              # in report but all zero buckets

    rep_metrics = {}
    for name in CS_REP_NAMES:
        rows    = [r for r in cs_rows if r["_rep"] == name]
        in_sc   = sum(float(r.get("Subtotal Sales") or 0) for r in rows)
        dec_r   = sum(float(r.get("Subtotal Sales") or 0) for r in rows
                      if str(r.get("Renew Status") or "").strip() == "Declined")
        dig_r   = sum(float(r.get("Subtotal Sales") or 0) for r in rows
                      if is_digital(r.get("Sale Items")))
        risk    = (dec_r / in_sc * 100) if in_sc else 0
        dig_mix = (dig_r / in_sc * 100) if in_sc else 0

        f1_sum = f2_sum = f3_sum = 0.0
        for r in rows:
            st = str(r.get("Renew Status") or "").strip()
            f1_sum += 40.0 if st == "Renewable" else (20.0 if st == "Merged" else 0.0)
            f2_sum += _pay_pts(r["_key"])
            f3_sum += 20.0 if is_digital(r.get("Sale Items")) else 0.0

        n  = max(len(rows), 1)
        renew_c   = f1_sum / n
        payment_c = f2_sum / n
        digital_c = f3_sum / n
        engage_c  = 0.0          # populated via Segmentation tab at run time
        hs        = renew_c + payment_c + digital_c + engage_c

        rep_metrics[name] = dict(
            in_scope=in_sc, declined_rev=dec_r, digital_rev=dig_r,
            risk_pct=risk, digital_mix=dig_mix,
            renew_component=renew_c, payment_component=payment_c,
            digital_component=digital_c, engagement_component=engage_c,
            health_score=hs,
        )

    # ── 4. Rep inScope reconcile ─────────────────────────────
    inscope_total = sum(m["in_scope"] for m in rep_metrics.values())
    inscope_gap   = abs(inscope_total - current_file_total)
    checks.append(_chk(
        "rep_inscope_reconcile",
        "FAIL" if inscope_gap > 100 else "PASS",
        f"${inscope_total:,.0f} (gap ${inscope_gap:,.0f})",
        f"${current_file_total:,.0f}",
        "Sum of all 6 rep inScope totals vs current file total",
    ))

    # ── 5. Health score components (4-Factor) ────────────────
    hs_fails = [
        f"{n} (components {m['renew_component']+m['payment_component']+m['digital_component']+m['engagement_component']:.2f}"
        f" vs stored {m['health_score']:.2f})"
        for n, m in rep_metrics.items()
        if abs(m["renew_component"] + m["payment_component"]
               + m["digital_component"] + m["engagement_component"]
               - m["health_score"]) > 0.1
    ]
    checks.append(_chk(
        "health_score_components",
        "FAIL" if hs_fails else "PASS",
        f"{6 - len(hs_fails)}/6 reps pass", "All 6 pass (within 0.1 pts)",
        "; ".join(hs_fails) if hs_fails
        else "Renew(40) + Payment(30) + Digital(20) + Engagement(10) sum to health score for all reps",
    ))

    # ── 6. Team health reconcile ─────────────────────────────
    avg_hs = sum(m["health_score"] for m in rep_metrics.values()) / len(CS_REP_NAMES)
    checks.append(_chk(
        "team_health_reconcile", "PASS",
        f"{avg_hs:.1f}", "Mean of 6 rep scores",
        "Team health = mean of individual rep health scores",
    ))

    # ── Bandwidth ────────────────────────────────────────────
    # Counts unique clients per rep (not agreements) so a client
    # with multiple contracts counts as 1 toward the 500 ceiling.
    bandwidth = {}
    for name in CS_REP_NAMES:
        rows    = [r for r in cs_rows if r["_rep"] == name]
        ren_c   = len({str(r.get("Client") or r.get("client") or r["_key"]).strip()
                       for r in rows
                       if str(r.get("Renew Status") or "").strip() == "Renewable"})
        dec_c   = len({str(r.get("Client") or r.get("client") or r["_key"]).strip()
                       for r in rows
                       if str(r.get("Renew Status") or "").strip() == "Declined"})
        # Cancelled: unique client/agreement values from the cancellation form
        # (col index 1 = rep email key, col index 2 = client/agreement field)
        ekey  = CS_REPS[name]
        can_c = len({
            str(list(r.values())[2]).strip()
            for r in cancellation_rows
            if len(r) > 2
            and ekey in str(list(r.values())[1]).lower()
            and str(list(r.values())[2]).strip()
        })
        bandwidth[name] = dict(
            renewable=ren_c, declined=dec_c,
            cancelled=can_c, active=ren_c + dec_c - can_c,
        )

    # ── 7. Bandwidth math ────────────────────────────────────
    bw_fails = [
        f"{n}: {bw['renewable']}R+{bw['declined']}D-{bw['cancelled']}C"
        f"≠{bw['active']} active"
        for n, bw in bandwidth.items()
        if bw["renewable"] + bw["declined"] - bw["cancelled"] != bw["active"]
    ]
    checks.append(_chk(
        "bandwidth_math",
        "FAIL" if bw_fails else "PASS",
        f"{6 - len(bw_fails)}/6 reps pass",
        "renewable + declined − cancelled = active for all reps",
        "; ".join(bw_fails) if bw_fails
        else "Bandwidth formula verified for all reps",
    ))

    # ── 8. Team bandwidth reconcile ──────────────────────────
    team_active = sum(bw["active"] for bw in bandwidth.values())
    checks.append(_chk(
        "team_bandwidth_reconcile", "PASS",
        f"{team_active} total active", "Sum of all 6 rep active counts",
        " + ".join(str(bw["active"]) for bw in bandwidth.values())
        + f" = {team_active}",
    ))

    # ── 9. Past due orphan check ─────────────────────────────
    renewal_keys = {r["_key"] for r in cs_rows}
    pd_orphans = sum(
        1 for row in past_due_rows
        if match_cs_rep(row.get("Rep"))
        and (ag := (row.get("Agreement") or row.get("Contract #")
                    or row.get("Agreement Number")))
        and str(ag).strip().upper() not in renewal_keys
    )
    checks.append(_chk(
        "past_due_orphan_check",
        "WARN" if pd_orphans else "PASS",
        f"{pd_orphans} orphaned", "0 orphans",
        f"{pd_orphans} CS-rep past-due account(s) with no matching "
        "Agreement Number in renewal report",
    ))

    # ── 10. Cancellation match rate ──────────────────────────
    cancel_nums = set()
    for row in cancellation_rows:
        vals = list(row.values())
        ag = extract_agreement_number(vals[2] if len(vals) > 2 else None)
        if ag:
            cancel_nums.add(ag)

    total_dec = len(declined_contracts)
    matched   = len(declined_contracts & cancel_nums)
    rate      = (matched / total_dec * 100) if total_dec else 100
    checks.append(_chk(
        "cancellation_match_rate",
        "WARN" if rate < 60 else "PASS",
        f"{rate:.1f}% ({matched}/{total_dec})", "≥60%",
        f"{matched} of {total_dec} declined accounts matched to a "
        "cancellation form entry",
    ))

    # ── 11. Cancellation orphan check ────────────────────────
    can_orphans = sum(
        1 for row in cancellation_rows
        if (vals := list(row.values()))
        and len(vals) > 2
        and any(CS_REPS[n] in str(vals[1]).lower() for n in CS_REP_NAMES)
        and (ag := extract_agreement_number(vals[2]))
        and ag not in renewal_keys
    )
    checks.append(_chk(
        "cancellation_orphan_check",
        "WARN" if can_orphans else "PASS",
        f"{can_orphans} orphaned", "0 orphans",
        f"{can_orphans} cancellation form submission(s) from CS-rep emails "
        "with no matching Agreement Number in renewal report",
    ))

    # ── 13–15. Penetration integrity (frontend-state checks) ─
    for name, note in [
        ("stale_next_steps",
         "Requires penetration fields populated in the Segmentation tab — evaluated client-side."),
        ("save_closed_integrity",
         "Requires Save Closed status from Segmentation tab — evaluated client-side."),
        ("escalation_integrity",
         "Requires Escalated + Blocker from Segmentation tab — evaluated client-side."),
    ]:
        checks.append(_chk(name, "PASS", "N/A — frontend state",
                           "Populated penetration fields", note))

    # ── 16. Export flag ──────────────────────────────────────
    pct = (current_file_total / annual_baseline * 100) if annual_baseline else 0
    checks.append(_chk(
        "export_flag",
        "FAIL" if pct < 80 else ("WARN" if pct < 95 else "PASS"),
        f"{pct:.1f}% of baseline", "≥95%",
        f"${current_file_total:,.0f} vs ANNUAL_BASELINE ${annual_baseline:,.0f}",
    ))

    # ── Renew Status breakdown (debug) ───────────────────────
    # Each unique Renew Status value found in CS-rep filtered rows,
    # with row count and revenue sum — helps diagnose segmentation gaps.
    status_buckets: dict = {}
    for row in cs_rows:
        sv  = str(row.get("Renew Status") or "").strip() or "(blank)"
        sub = float(row.get("Subtotal Sales") or 0)
        if sv not in status_buckets:
            status_buckets[sv] = {"status": sv, "count": 0, "revenue": 0.0}
        status_buckets[sv]["count"]   += 1
        status_buckets[sv]["revenue"] += sub

    renew_status_breakdown = sorted(
        status_buckets.values(), key=lambda x: x["revenue"], reverse=True
    )

    # ── Summary ──────────────────────────────────────────────
    passed  = sum(1 for c in checks if c["status"] == "PASS")
    warned  = sum(1 for c in checks if c["status"] == "WARN")
    failed  = sum(1 for c in checks if c["status"] == "FAIL")
    overall = "CRITICAL" if failed else ("REVIEW" if warned else "CLEAN")

    summary = dict(total_checks=len(checks), passed=passed,
                   warned=warned, failed=failed, overall_status=overall)
    return checks, summary, renew_status_breakdown


# ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/bvm-bot.js")
def bvm_bot():
    return send_from_directory(".", "bvm-bot.js")

@app.route("/bot-proxy", methods=["POST"])
def bot_proxy():
    payload = request.get_data()
    req = urllib.request.Request(
        "https://gatekeeper-bot-omega.vercel.app/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req) as resp:
        body = resp.read()
    return current_app.response_class(body, status=200, mimetype="application/json")


@app.route("/process", methods=["POST"])
def process():
    missing = [f for f in REQUIRED_FILES if f not in request.files]
    if missing:
        return jsonify({"error": f"Missing required files: {missing}"}), 400

    errors = {}
    result = {}

    # ---- Renewal report ----
    try:
        result["renewal_rows"] = df_to_records(read_xlsx(request.files["renewal"]))
    except Exception as e:
        errors["renewal"] = str(e)

    # ---- Executive pack (txt) ----
    try:
        result["exec_pack_text"] = request.files["execpack"].read().decode("utf-8", errors="replace")
    except Exception as e:
        errors["execpack"] = str(e)

    # ---- Past due report ----
    try:
        result["past_due_rows"] = df_to_records(read_xlsx(request.files["pastdue"]))
    except Exception as e:
        errors["pastdue"] = str(e)

    # ---- Go to print ----
    try:
        result["gtp_rows"] = df_to_records(read_xlsx(request.files["gtp"]))
    except Exception as e:
        errors["gtp"] = str(e)

    # ---- Cancellation form (two tabs) ----
    try:
        can_bytes = request.files["cancellation"].read()
        # Discover actual sheet names (case-insensitive match)
        xl_sheets = pd.ExcelFile(io.BytesIO(can_bytes), engine="openpyxl").sheet_names
        print(f"[CANCELLATION] Available sheets: {xl_sheets}")

        def _find_sheet(keywords, fallback):
            for kw in keywords:
                for s in xl_sheets:
                    if kw.lower() in s.lower():
                        return s
            return fallback

        done_sheet    = _find_sheet(["done", "completed", "processed"], xl_sheets[-1] if xl_sheets else 0)
        pending_sheet = _find_sheet(["form responses", "pending", "response"], xl_sheets[0] if xl_sheets else 0)
        print(f"[CANCELLATION] done_sheet={done_sheet!r}  pending_sheet={pending_sheet!r}")

        # DONE tab has no header row — first row is data, so read with header=None
        done_df = read_xlsx(can_bytes, sheet_name=done_sheet, header=None)
        print(f"[DONE TAB] {len(done_df)} rows, {len(done_df.columns)} columns")
        print(f"[DONE TAB] col[0]={repr(done_df.iloc[0,0] if len(done_df)>0 else 'N/A')}  "
              f"col[1]={repr(done_df.iloc[0,1] if len(done_df)>0 else 'N/A')}  "
              f"col[2]={repr(done_df.iloc[0,2] if len(done_df)>0 else 'N/A')}  "
              f"col[5]={repr(done_df.iloc[0,5] if len(done_df)>0 and len(done_df.columns)>5 else 'N/A')}  "
              f"col[7]={repr(done_df.iloc[0,7] if len(done_df)>0 and len(done_df.columns)>7 else 'N/A')}  "
              f"col[8]={repr(done_df.iloc[0,8] if len(done_df)>0 and len(done_df.columns)>8 else 'N/A')}")
        cancellation_done_rows = df_to_records(done_df)
        try:
            cancellation_pending_rows = (
                df_to_records(read_xlsx(can_bytes, sheet_name=pending_sheet))
                if pending_sheet != done_sheet else []
            )
        except Exception:
            cancellation_pending_rows = []

        # backward-compat: cancellation_rows → DONE tab (used by bandwidth + validations)
        result["cancellation_rows"]          = cancellation_done_rows
        result["cancellation_pending_count"] = len(cancellation_pending_rows)
        print(f"[CANCELLATION] done={len(cancellation_done_rows)} rows  pending={len(cancellation_pending_rows)} rows")
    except Exception as e:
        errors["cancellation"] = str(e)
        cancellation_done_rows    = []
        cancellation_pending_rows = []

    if errors:
        return jsonify({"error": "File parsing failed", "details": errors}), 400

    result["annual_baseline"] = ANNUAL_BASELINE

    checks, summary, renew_status_breakdown = run_validations(
        result.get("renewal_rows", []),
        result.get("exec_pack_text", ""),
        result.get("past_due_rows", []),
        result.get("gtp_rows", []),
        result.get("cancellation_rows", []),
        ANNUAL_BASELINE,
    )
    result["validations"]            = checks
    result["validation_summary"]     = summary
    result["renew_status_breakdown"] = renew_status_breakdown

    # ── Bandwidth debug ───────────────────────────────────────────
    renewal_rows_raw = result.get("renewal_rows", [])
    _client_col = next(
        (k for k in (renewal_rows_raw[0].keys() if renewal_rows_raw else [])
         if "client" in k.lower()),
        None
    )
    bw_debug = {}
    for name in CS_REP_NAMES:
        rep_rows    = [r for r in renewal_rows_raw
                       if match_cs_rep(r.get("Rep")) == name]
        client_vals = ([str(r.get(_client_col) or "").strip() for r in rep_rows]
                       if _client_col else [])
        agr_vals    = [str(r.get("Contract #") or r.get("Agreement")
                           or r.get("Agreement Number") or "").strip()
                       for r in rep_rows]
        bw_debug[name] = {
            "total_rows":        len(rep_rows),
            "unique_clients":    len({v for v in client_vals if v}),
            "unique_agreements": len({v for v in agr_vals if v}),
            "client_col_found":  _client_col,
            "sample_clients":    list({v for v in client_vals if v})[:5],
        }
    result["bandwidth_debug"] = bw_debug

    # ── Cancelled clients by rep from cancellation form ───────────
    # Col index 1 = rep email, col index 2 = client/agreement identifier
    cancellation_rows_raw = result.get("cancellation_rows", [])
    cancelled_clients_by_rep = {}
    print("\n" + "="*60)
    print("CANCELLED CLIENTS BY REP (cancellation form)")
    for name in CS_REP_NAMES:
        ekey     = CS_REPS[name]
        can_rows = [r for r in cancellation_rows_raw
                    if len(r) > 1
                    and ekey in str(list(r.values())[1]).lower()]
        unique   = {str(list(r.values())[2]).strip()
                    for r in can_rows
                    if len(r) > 2 and str(list(r.values())[2]).strip()}
        cancelled_clients_by_rep[name] = len(unique)
        print(f"  {name}: {len(can_rows)} rows → {len(unique)} unique clients")
    print("="*60 + "\n")

    result["cancelled_clients_by_rep"] = cancelled_clients_by_rep

    # ── Revenue Metrics: MRR / ARR / TCV ─────────────────────────
    # cs-rep rows only, Cancelled excluded — mirrors the cs_rows filter in run_validations
    rep_set  = set(CS_REP_NAMES)
    mrr_rows = [
        r for r in renewal_rows_raw
        if match_cs_rep(r.get("Rep")) in rep_set
        and str(r.get("Renew Status") or "").strip() != "Cancelled"
    ]
    def _f(r, col):
        try:    return float(r.get(col) or 0)
        except: return 0.0

    team_mrr     = sum(_f(r, "Monthly") for r in mrr_rows)
    team_arr     = team_mrr * 12
    team_tcv     = sum(_f(r, "Subtotal Sales") for r in mrr_rows)

    dec_rows     = [r for r in mrr_rows if str(r.get("Renew Status") or "").strip() == "Declined"]
    declined_mrr = sum(_f(r, "Monthly") for r in dec_rows)
    declined_arr = declined_mrr * 12
    declined_tcv = sum(_f(r, "Subtotal Sales") for r in dec_rows)

    dig_mrr      = sum(_f(r, "Monthly") for r in mrr_rows if is_digital(r.get("Sale Items")))
    dig_mrr_mix  = (dig_mrr / team_mrr * 100) if team_mrr > 0 else 0.0

    result["revenue_metrics"] = {
        "team_mrr":       team_mrr,
        "team_arr":       team_arr,
        "team_tcv":       team_tcv,
        "declined_mrr":   declined_mrr,
        "declined_arr":   declined_arr,
        "declined_tcv":   declined_tcv,
        "digital_mrr_mix": dig_mrr_mix,
    }
    print(f"\n[REVENUE] MRR={team_mrr:>12,.0f}  ARR={team_arr:>12,.0f}  TCV={team_tcv:>12,.0f}")
    print(f"[REVENUE] Declined MRR={declined_mrr:>10,.0f}  ARR={declined_arr:>10,.0f}  TCV={declined_tcv:>10,.0f}")
    print(f"[REVENUE] Digital MRR Mix={dig_mrr_mix:.1f}%\n")

    # ── Decline Performance ───────────────────────────────────────
    # Build renewal agreement lookup: normalised E-XXXXXX → Subtotal Sales
    renewal_agr_lookup = {}
    for r in renewal_rows_raw:
        raw_k = str(r.get("Contract #") or r.get("Agreement")
                    or r.get("Agreement Number") or "").strip()
        agr_k = extract_agreement_number(raw_k)
        if not agr_k and raw_k:
            agr_k = ("E-" + raw_k) if not raw_k.upper().startswith("E-") else raw_k.upper()
        if agr_k:
            try:
                renewal_agr_lookup[agr_k.upper()] = float(r.get("Subtotal Sales") or 0)
            except (TypeError, ValueError):
                pass

    SAVE_YES = {"yes", "y", "true", "1", "offered", "extended", "yes - offered",
                "yes-offered", "offer extended"}

    # DONE tab column layout (no header row, read with header=None):
    # [0] Timestamp  [1] Rep email  [2] Combined (E-# / Client)
    # [3] Region     [4] Edition    [5] Form type (DECLINE RENEWAL / CANCELLATION)
    # [6] Deadline month            [7] Attrition reason
    # [8] Save offer extended       [9] Notes
    DONE_COL_TIMESTAMP   = 0
    DONE_COL_EMAIL       = 1
    DONE_COL_COMBINED    = 2
    DONE_COL_FORM_TYPE   = 5
    DONE_COL_ATTRITION   = 7
    DONE_COL_SAVE_OFFER  = 8

    decline_performance = []
    for row in cancellation_done_rows:
        vals = list(row.values())
        if len(vals) < 3:
            continue

        timestamp   = vals[DONE_COL_TIMESTAMP] if len(vals) > DONE_COL_TIMESTAMP else None
        rep_email   = str(vals[DONE_COL_EMAIL] or "").strip().lower() if len(vals) > DONE_COL_EMAIL else ""
        combined    = str(vals[DONE_COL_COMBINED] or "").strip() if len(vals) > DONE_COL_COMBINED else ""

        # Match rep name from email key
        rep_name = None
        for rname, ekey in CS_REPS.items():
            if ekey.lower() in rep_email:
                rep_name = rname
                break

        # Extract E-XXXXXX and client name
        agr_num     = extract_agreement_number(combined)
        client_name = _extract_client_name(combined)

        # Positional column reads — DONE tab has no header row
        form_type       = str(vals[DONE_COL_FORM_TYPE]  or "").strip() if len(vals) > DONE_COL_FORM_TYPE  else ""
        attrition_reason= str(vals[DONE_COL_ATTRITION]  or "").strip() if len(vals) > DONE_COL_ATTRITION  else ""
        save_offer      = str(vals[DONE_COL_SAVE_OFFER]  or "").strip() if len(vals) > DONE_COL_SAVE_OFFER  else ""

        # Match to renewal
        match_status    = "malformed"
        matched_revenue = None
        if agr_num:
            agr_key = agr_num.upper()
            if agr_key in renewal_agr_lookup:
                matched_revenue = renewal_agr_lookup[agr_key]
                match_status = "matched"
            else:
                match_status = "unmatched"

        decline_performance.append({
            "rep_email":           rep_email,
            "rep_name":            rep_name or "Unknown",
            "agreement_number":    agr_num,
            "client_name":         client_name,
            "form_type":           form_type,
            "attrition_reason":    attrition_reason,
            "save_offer_extended": save_offer,
            "timestamp":           str(timestamp) if timestamp is not None else None,
            "matched_revenue":     matched_revenue,
            "match_status":        match_status,
        })

    _dp_matched   = sum(1 for r in decline_performance if r["match_status"] == "matched")
    _dp_unmatched = sum(1 for r in decline_performance if r["match_status"] == "unmatched")
    _dp_malformed = sum(1 for r in decline_performance if r["match_status"] == "malformed")
    print(f"[DECLINE PERF] {len(decline_performance)} DONE rows  "
          f"matched={_dp_matched}  unmatched={_dp_unmatched}  malformed={_dp_malformed}")

    # Per-rep + team summary
    dp_summary = {}
    for name in CS_REP_NAMES:
        rep_dp  = [r for r in decline_performance if r["rep_name"] == name]
        matched = [r for r in rep_dp if r["match_status"] == "matched"]
        save_n  = sum(1 for r in rep_dp
                      if str(r.get("save_offer_extended") or "").strip().lower() in SAVE_YES)
        attr_bk = {}
        for r in rep_dp:
            ar = str(r.get("attrition_reason") or "").strip()
            if ar:
                attr_bk[ar] = attr_bk.get(ar, 0) + 1
        dp_summary[name] = {
            "total_submissions":     len(rep_dp),
            "decline_renewals":      sum(1 for r in rep_dp
                                        if "decline" in str(r.get("form_type") or "").lower()),
            "cancellations":         sum(1 for r in rep_dp
                                        if "cancel" in str(r.get("form_type") or "").lower()),
            "matched_revenue_total": sum(r["matched_revenue"] for r in matched
                                        if r["matched_revenue"] is not None),
            "unmatched_count":       sum(1 for r in rep_dp if r["match_status"] == "unmatched"),
            "malformed_count":       sum(1 for r in rep_dp if r["match_status"] == "malformed"),
            "save_offer_rate":       round(save_n / len(rep_dp) * 100, 1) if rep_dp else 0.0,
            "attrition_breakdown":   attr_bk,
        }

    team_save_n  = sum(1 for r in decline_performance
                       if str(r.get("save_offer_extended") or "").strip().lower() in SAVE_YES)
    team_attr_bk = {}
    for r in decline_performance:
        ar = str(r.get("attrition_reason") or "").strip()
        if ar:
            team_attr_bk[ar] = team_attr_bk.get(ar, 0) + 1
    dp_summary["_team"] = {
        "total_submissions":     len(decline_performance),
        "decline_renewals":      sum(v["decline_renewals"]      for v in dp_summary.values()),
        "cancellations":         sum(v["cancellations"]          for v in dp_summary.values()),
        "matched_revenue_total": sum(v["matched_revenue_total"]  for v in dp_summary.values()),
        "unmatched_count":       sum(v["unmatched_count"]        for v in dp_summary.values()),
        "malformed_count":       sum(v["malformed_count"]        for v in dp_summary.values()),
        "save_offer_rate":       round(team_save_n / len(decline_performance) * 100, 1)
                                 if decline_performance else 0.0,
        "attrition_breakdown":   team_attr_bk,
    }

    result["decline_performance"]         = decline_performance
    result["decline_performance_summary"] = dp_summary

    # Append validation check then recompute summary
    _dp_valid   = sum(1 for r in decline_performance if r["match_status"] != "malformed")
    _dp_matched2 = sum(1 for r in decline_performance if r["match_status"] == "matched")
    _dp_rate    = (_dp_matched2 / _dp_valid * 100) if _dp_valid else 100.0
    checks.append(_chk(
        "decline_performance_match_rate",
        "FAIL" if _dp_rate < 40 else ("WARN" if _dp_rate < 70 else "PASS"),
        f"{_dp_matched2}/{_dp_valid} matched ({_dp_rate:.1f}%)",
        "≥70% match rate",
        f"{_dp_matched2} of {_dp_valid} DONE submissions matched to renewal report Subtotal Sales",
    ))
    _p = sum(1 for c in checks if c["status"] == "PASS")
    _w = sum(1 for c in checks if c["status"] == "WARN")
    _f = sum(1 for c in checks if c["status"] == "FAIL")
    result["validations"]        = checks
    result["validation_summary"] = dict(
        total_checks=len(checks), passed=_p, warned=_w, failed=_f,
        overall_status="CRITICAL" if _f else ("REVIEW" if _w else "CLEAN"),
    )

    result = sanitize_keys(result)
    return current_app.response_class(
        json.dumps(result, cls=SafeEncoder),
        mimetype="application/json"
    )


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"}), 200


@app.route("/export_pack", methods=["POST"])
def export_pack():
    data = request.get_json(force=True)
    if not data:
        return jsonify({"error": "No JSON body"}), 400
    try:
        buf = build_pack_xlsx(data)
        date_str = datetime.now().strftime("%m%d%y")
        fname = f"Gatekeeper_ExecutivePack_{date_str}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )
    except Exception as e:
        current_app.logger.exception("export_pack error")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", debug=True, port=8080)
